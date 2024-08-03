from flask import Flask, render_template, Response, request, jsonify, url_for, redirect
import datetime
import os
import cv2
import numpy as np
from PIL import Image
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# In-memory storage for records
records = []

segments = {
    (1, 1, 1, 0, 1, 1, 1): 0,
    (0, 0, 1, 0, 0, 1, 0): 1,
    (1, 0, 1, 1, 1, 0, 1): 2,
    (1, 0, 1, 1, 0, 1, 1): 3,
    (0, 1, 1, 1, 0, 1, 0): 4,
    (1, 1, 0, 1, 0, 1, 1): 5,
    (1, 1, 0, 1, 1, 1, 1): 6,
    (1, 0, 1, 0, 0, 1, 0): 7,
    (1, 1, 1, 1, 1, 1, 1): 8,
    (1, 1, 1, 1, 0, 1, 1): 9
}

EXCEL_FILE = 'consumption_records.xlsx'

def initialize_excel_file():
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Consumption Records"
        sheet.append(["Timestamp", "Consumption (KWh)", "Carbon Footprint (Kg CO2)"])
        workbook.save(EXCEL_FILE)

def append_to_excel(record):
    if not os.path.exists(EXCEL_FILE):
        initialize_excel_file()
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([record['timestamp'], record['consumption'], record['carbon_footprint']])
    workbook.save(EXCEL_FILE)

def get_digit(img, segment_pos):
    active = map(lambda x: int(np.count_nonzero(get_dig_sub(img, x[0], x[1], 4, 4)) > 8), segment_pos)
    return segments.get(tuple(active), 'x')

def get_dig_sub(img, x, y, width, height):
    return img[y:y + height, x:x + width]

def extract_and_identify_digit(img, x, y, width, height, segment_pos):
    imgray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    ret, imthresh = cv2.threshold(imgray, 76, 255, cv2.THRESH_BINARY_INV)
    imthresh = cv2.dilate(imthresh, np.ones((2, 2), np.uint8), iterations=5)
    digit_area = get_dig_sub(imthresh, x, y, width, height)
    digit = get_digit(digit_area, segment_pos)
    return digit

def capture_image_from_webcam(save_path='captured_image.png'):
    cap = cv2.VideoCapture(0)  
    if not cap.isOpened():
        return None
    ret, frame = cap.read()
    cap.release()
    if not ret:
        return None
    cv2.imwrite(save_path, frame)  
    return frame

def generate_frames():
    cap = cv2.VideoCapture(0)
    while True:
        success, frame = cap.read()
        if not success:
            break
        else:
            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

@app.route('/')
def index():
    remark = records[-1]['remark'] if records else ''  # Get the remark from the latest record if available
    return render_template('index.html', records=records, remark=remark)

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/capture', methods=['POST'])
def capture():
    image = capture_image_from_webcam()
    if image is None:
        return jsonify({'error': 'Failed to capture image from webcam'})

    digit_parameters = [
        (0, 227, 96, 186, [(58, 15), (17, 53), (89, 54), (48, 95), (10, 132), (78, 138), (38, 174)]),  # first digit
        (100, 227, 100, 186, [(61, 16), (22, 59), (94, 67), (54, 98), (13, 143), (86, 147), (44, 178)]),  # second digit
        (212, 230, 100, 186, [(40, 9), (16, 38), (81, 53), (45, 87), (8, 119), (70, 142), (42, 171)]),  # third digit
        (325, 285, 79, 135, [(44, 19), (6, 45), (65, 50), (34, 73), (8, 108), (62, 107), (31, 127)])  # fourth digit
    ]

    identified_digits = []
    for (x, y, width, height, segment_pos) in digit_parameters:
        identified_digit = extract_and_identify_digit(image, x, y, width, height, segment_pos)
        identified_digits.append(str(identified_digit))
    
    result_string = ''.join(identified_digits[:3]) + '.' + identified_digits[3] + ' KWh'
    consumption_value = float(result_string.split()[0])
    carbon_footprint = consumption_value * 0.85

    record = {
        'timestamp': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'consumption': result_string,
        'carbon_footprint': round(carbon_footprint, 2),
        'remark': 'Good Work!' if carbon_footprint < 7085 else 'You need to reduce your consumption'
    }
    records.append(record)
    append_to_excel(record)
    return jsonify(record)

@app.route('/graph')
def graph():
    timestamps = [record['timestamp'] for record in records]
    carbon_footprints = [record['carbon_footprint'] for record in records]

    plt.figure(figsize=(10, 5))
    plt.plot(timestamps, carbon_footprints, marker='o')
    plt.xlabel('Timestamp')
    plt.ylabel('Carbon Footprint (Kg CO2)')
    plt.title('Carbon Footprint Over Time')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    
    graph_path = os.path.join('static', 'graph.png')
    plt.savefig(graph_path)
    return redirect(url_for('static', filename='graph.png'))

if __name__ == "__main__":
    app.run(debug=True)
